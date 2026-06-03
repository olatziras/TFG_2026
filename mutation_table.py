import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from Levenshtein import editops, distance

# ==========================================
# 1. CONFIGURACIÓN (Pon tus datos aquí)
# ==========================================
spacer_ref = "CCGATGGGATCCGACGTTCGCCGTCTGCGTTACC"  # Pega aquí tu secuencia de interés
archivo_fastq = "8S3ZPF_1_769_PCR2_2.fastq"             # Nombre exacto de tu archivo FASTQ
max_errores = 5                                 # Filtro estricto de Levenshtein
# ==========================================

longitud = len(spacer_ref)
# Creamos la lista de "cajitas" para acumular mutaciones por posición
conteo_mutaciones = [0] * longitud

print(f"Iniciando escaneo... Buscando el spacer (máx {max_errores} mutaciones)...")

# 2. PROCESAMIENTO DEL ARCHIVO FASTQ
with open(archivo_fastq, "r") as f:
    for i, line in enumerate(f):
        if i % 4 == 1:  # Extraer solo las líneas de secuencia de ADN
            read = line.strip().upper()
            
            # Ventana deslizante para cazar la mejor coincidencia de Levenshtein
            mejor_dist = max_errores + 1
            mejor_subseq = ""
            for j in range(len(read) - longitud + 1):
                subseq = read[j:j+longitud]
                d = distance(spacer_ref, subseq)
                if d < mejor_dist:
                    mejor_dist = d
                    mejor_subseq = subseq
            
            # Si el fragmento pasa el filtro de calidad (máximo 3 errores)
            if mejor_dist <= max_errores:
                ops = editops(spacer_ref, mejor_subseq)
                # Extraemos la posición exacta del error (s_idx)
                for op, s_idx, d_idx in ops:
                    if s_idx < longitud:
                        conteo_mutaciones[s_idx] += 1

print("Escaneo completado. Diseñando tu archivo Excel...")

# 3. CREACIÓN Y DISEÑO DEL EXCEL PROFESIONAL
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Perfil de Mutaciones"
ws.views.sheetView[0].showGridLines = True  # Asegura que las cuadrículas se vean

# Estilos visuales estables (Gama Gris Azulado)
HEADER_BG = "2C4A5E"
ZEBRA_BG = "F4F7F9"
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# Títulos del documento
ws['A1'] = "ANÁLISIS DE VARIABILIDAD POR POSICIÓN"
ws['A1'].font = Font(name="Segoe UI", size=16, bold=True, color="2C4A5E")
ws['A2'] = f"Referencia: {spacer_ref} (Margen de Levenshtein: Máx {max_errores} mutaciones)"
ws['A2'].font = Font(name="Segoe UI", size=10, italic=True, color="555555")

# Escribir Cabeceras
cabeceras = ["Posición", "Base Original", "Conteo de Mutaciones"]
for col_idx, text in enumerate(cabeceras, start=1):
    cell = ws.cell(row=4, column=col_idx, value=text)
    cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
ws.row_dimensions[4].height = 26

# Rellenar filas con los datos recolectados
start_row = 5
for i in range(longitud):
    row_idx = start_row + i
    
    c_pos = ws.cell(row=row_idx, column=1, value=i + 1)
    c_base = ws.cell(row=row_idx, column=2, value=spacer_ref[i])
    c_count = ws.cell(row=row_idx, column=3, value=conteo_mutaciones[i])
    
    # Formatos y Alineaciones
    c_pos.alignment = Alignment(horizontal="center")
    c_base.alignment = Alignment(horizontal="center")
    c_count.alignment = Alignment(horizontal="right")
    c_count.number_format = '#,##0'
    
    c_pos.font = Font(name="Segoe UI", size=11)
    c_base.font = Font(name="Consolas", size=11, bold=True, color="1E3A8A")  # Monoespaciado para letras de ADN
    c_count.font = Font(name="Segoe UI", size=11)
    
    # Zebra striping (Filas alternas claras para mejorar la lectura)
    for cell in (c_pos, c_base, c_count):
        cell.border = thin_border
        if i % 2 == 1:
            cell.fill = PatternFill(start_color=ZEBRA_BG, end_color=ZEBRA_BG, fill_type="solid")
            
    ws.row_dimensions[row_idx].height = 20

# Añadir fila de totales abajo
total_row = start_row + longitud
ws.cell(row=total_row, column=1, value="TOTAL MUTACIONES ACUMULADAS").font = Font(name="Segoe UI", size=11, bold=True)
ws.cell(row=total_row, column=1).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
ws.cell(row=total_row, column=2).border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))

c_total = ws.cell(row=total_row, column=3, value=f"=SUM(C{start_row}:C{total_row-1})")
c_total.font = Font(name="Segoe UI", size=11, bold=True)
c_total.alignment = Alignment(horizontal="right")
c_total.number_format = '#,##0'
c_total.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
ws.row_dimensions[total_row].height = 24

# Autoajustar el tamaño de las columnas
for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    if col[0].column <= 3:
        ws.column_dimensions[col_letter].width = 24

# Guardar Archivo Final
nombre_salida = "Resultado_Perfil_Mutaciones.xlsx"
wb.save(nombre_salida)
print(f"¡Éxito total! Archivo Excel guardado como: {nombre_salida}")
